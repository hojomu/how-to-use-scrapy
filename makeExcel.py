import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook

def get_json_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        return json.load(file)
    
def json_to_df(data, columns):
    df = pd.DataFrame(data, columns=columns)
    df = df.astype(str)
    return df

def change_workbook(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    
    # ws.auto_filter.ref = 'A1:B{}'.format(ws.max_row) # A, B 열 각각

    header_fill = PatternFill(start_color="90EE90", fill_type="solid")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    ws.freeze_panes = 'A2' # 상단 행 고정

    tab = Table(displayName="Table1", ref=ws.dimensions)
    ws.add_table(tab)

    wb.save(file_path)
    

# 기본 값 세팅
subject = 'robotworld'

json_file_path = f'C:/pythonproject/ScrapyFW/ScrapyFW/results/{subject}/{subject}.json'
excel_file_path = f'C:/pythonproject/ScrapyFW/ScrapyFW/results/{subject}/{subject}.xlsx'

json_data = get_json_file(json_file_path)


columns = [
    'name',
    'email',
    'homepage',
    'category',
    'category_detail',
    'product_introduce',
    ]

df = json_to_df(json_data, columns)
df.columns = ['업체명', '이메일', '홈페이지', '카테고리','카테고리 상세','제품 소개']
df.to_excel(excel_file_path, index=False, engine='openpyxl')

change_workbook(excel_file_path)

print(f"Excel file saved to {excel_file_path}")